"""
Consulta de produtos em lote.

Dois blocos:
  1. Extração de EANs de fontes diversas (texto colado, PDF, CSV, XLSX).
  2. Lookup paralelo (ThreadPoolExecutor) reaproveitando o EANPicturesService,
     com cache e cascata já embutidos.

A extração é tolerante: varre números de 8/12/13/14 dígitos e mantém apenas os
que passam no dígito verificador (is_valid_ean), descartando lixo de OCR/colunas.
"""

from __future__ import annotations

import csv
import io
import re
from concurrent.futures import ThreadPoolExecutor

from .ean_service import (
    APIUnavailableError,
    EANPicturesService,
    InvalidEANError,
    ProductNotFoundError,
)

# Sequências de 8 a 14 dígitos (com possíveis separadores . - espaço no meio
# são removidos antes). \b evita capturar pedaços de números maiores.
_EAN_RE = re.compile(r"(?<!\d)(\d{8}|\d{12,14})(?!\d)")

# Limites defensivos para uploads.
MAX_EANS = 1000
COSMOS_PRODUCTS_BASE_URL = "https://cdn-cosmos.bluesoft.com.br/products"


# --------------------------------------------------------------------------- #
# Extração de EANs
# --------------------------------------------------------------------------- #
def extract_eans_from_text(text: str) -> list[str]:
    """
    Extrai EANs válidos de um texto livre, preservando a ordem e sem duplicar.

    Remove separadores comuns (espaço, ponto, hífen) entre dígitos antes de casar.
    """
    if not text:
        return []
    # Junta dígitos separados por . ou - (códigos formatados). NÃO removemos
    # espaços/quebras de linha de propósito: isso fundiria EANs adjacentes
    # (ex.: dois códigos em linhas seguidas) num número único e inválido.
    normalized = re.sub(r"(?<=\d)[.\-](?=\d)", "", text)
    seen: set[str] = set()
    out: list[str] = []
    for match in _EAN_RE.findall(normalized):
        if match not in seen and EANPicturesService.is_valid_ean(match):
            seen.add(match)
            out.append(match)
    return out


def extract_eans_from_pdf(data: bytes) -> list[str]:
    """Extrai EANs de um PDF (texto). Requer pypdf instalado."""
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise RuntimeError(
            "Leitura de PDF requer o pacote 'pypdf' (pip install pypdf)."
        ) from exc

    reader = PdfReader(io.BytesIO(data))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    return extract_eans_from_text(text)


def build_cosmos_product_image_url(ean: str) -> str:
    """Monta a URL direta da imagem de um produto na CDN da Cosmos."""
    normalized = EANPicturesService.normalize_ean(ean)
    if not EANPicturesService.is_valid_ean(normalized):
        raise InvalidEANError(f"Código EAN inválido: {normalized!r}")
    return f"{COSMOS_PRODUCTS_BASE_URL}/{normalized}"


def extract_cosmos_image_urls_from_pdf(data: bytes) -> list[str]:
    """Extrai EANs válidos de um PDF e gera uma URL Cosmos para cada um."""
    return [build_cosmos_product_image_url(ean) for ean in extract_eans_from_pdf(data)]


def extract_eans_from_csv(data: bytes) -> list[str]:
    """Extrai EANs de um CSV (qualquer coluna)."""
    text = data.decode("utf-8-sig", errors="replace")
    rows = csv.reader(io.StringIO(text))
    return extract_eans_from_text("\n".join(" ".join(row) for row in rows))


def extract_eans_from_xlsx(data: bytes) -> list[str]:
    """Extrai EANs de um XLSX (todas as abas/células). Requer openpyxl."""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise RuntimeError(
            "Leitura de XLSX requer o pacote 'openpyxl' (pip install openpyxl)."
        ) from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            parts.append(" ".join(str(c) for c in row if c is not None))
    wb.close()
    return extract_eans_from_text("\n".join(parts))


def extract_eans_from_file(filename: str, data: bytes) -> list[str]:
    """Despacha a extração pelo tipo do arquivo (extensão)."""
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        return extract_eans_from_pdf(data)
    if name.endswith(".csv"):
        return extract_eans_from_csv(data)
    if name.endswith((".xlsx", ".xlsm")):
        return extract_eans_from_xlsx(data)
    # txt, json ou desconhecido: trata como texto puro.
    return extract_eans_from_text(data.decode("utf-8", errors="replace"))


# --------------------------------------------------------------------------- #
# Lookup em lote
# --------------------------------------------------------------------------- #
def _lookup_one(service: EANPicturesService, ean: str, providers: list[str] | None) -> dict:
    """Consulta um EAN e devolve uma linha padronizada (nunca levanta)."""
    try:
        p = service.get_product(ean, providers=providers)
        return {
            "ean": p.get("ean", ean),
            "found": True,
            "name": p.get("name", ""),
            "image": p.get("image", ""),
            "description": p.get("description", ""),
            "source": p.get("source", ""),
            "from_cache": bool(p.get("from_cache")),
            "error": "",
            "code": "",
        }
    except InvalidEANError as exc:
        return _row_error(ean, "invalid_ean", str(exc))
    except ProductNotFoundError as exc:
        return _row_error(ean, "not_found", str(exc))
    except APIUnavailableError as exc:
        return _row_error(ean, "api_unavailable", str(exc))


def _row_error(ean: str, code: str, message: str) -> dict:
    return {
        "ean": ean,
        "found": False,
        "name": "",
        "image": "",
        "description": "",
        "source": "",
        "from_cache": False,
        "error": message,
        "code": code,
    }


def lookup_batch(
    service: EANPicturesService,
    eans: list[str],
    providers: list[str] | None = None,
    max_workers: int = 8,
) -> list[dict]:
    """
    Consulta vários EANs em paralelo, preservando a ordem de entrada.

    Resultados saem como dicts com: ean, found, name, image, description,
    source, from_cache, error, code. Limitado a MAX_EANS por chamada.
    """
    eans = eans[:MAX_EANS]
    if not eans:
        return []

    workers = max(1, min(max_workers, len(eans)))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        # map preserva a ordem da lista de entrada.
        return list(pool.map(lambda e: _lookup_one(service, e, providers), eans))


def summarize(results: list[dict]) -> dict:
    """Resumo agregado de um lote (para exibir no front)."""
    total = len(results)
    found = sum(1 for r in results if r["found"])
    with_image = sum(1 for r in results if r["found"] and r["image"])
    return {
        "total": total,
        "found": found,
        "not_found": total - found,
        "with_image": with_image,
    }


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
_EXPORT_COLUMNS = ["ean", "name", "image", "description", "source", "found", "error"]


def results_to_csv(results: list[dict]) -> str:
    """Serializa os resultados em CSV (UTF-8)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for r in results:
        writer.writerow({**r, "found": "sim" if r.get("found") else "não"})
    return buf.getvalue()


def results_to_xlsx(results: list[dict]) -> bytes:
    """Serializa os resultados em XLSX (bytes). Requer openpyxl."""
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise RuntimeError(
            "Export XLSX requer o pacote 'openpyxl' (pip install openpyxl)."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    ws.append(_EXPORT_COLUMNS)
    for r in results:
        ws.append([
            r.get("ean", ""),
            r.get("name", ""),
            r.get("image", ""),
            r.get("description", ""),
            r.get("source", ""),
            "sim" if r.get("found") else "não",
            r.get("error", ""),
        ])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
